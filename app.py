import openpyxl
import streamlit as st
import pickle
from streamlit_option_menu import option_menu
model=pickle.load(open('model.pkl','rb'))
st.title("Multi-SMS Classifier")
text=['']
v={2:'Fund Transfer',0: 'Login OTP',1: 'Password reset OTP',4: 'SPAM',3: 'HAM'}
if st.button("Classify Your Messsages"):
    wb=openpyxl.load_workbook('received messages.xlsx')
    ws=wb.active
    wb_pr=openpyxl.load_workbook('messages.xlsx')
    ws_pr=wb_pr.active
    sms_count=1
    sms_idx='A'+str(sms_count)
    sms=ws[sms_idx].value
    if sms:
        while sms:
            text[0]=sms
            result=model.predict(text)
            if result[0]==0:
                for i in range(1,1000):
                    idx='C'+str(i)
                    if ws_pr[idx].value==None:
                        ws_pr[idx].value=text[0]
                        break
            elif result[0]==1:
                for i in range(1,1000):
                    idx='D'+str(i)
                    if ws_pr[idx].value==None:
                        ws_pr[idx].value=text[0]
                        break
            elif result[0]==2:
                for i in range(1,1000):
                    idx='B'+str(i)
                    if ws_pr[idx].value==None:
                        ws_pr[idx].value=text[0]
                        break
            elif result[0]==3:
                for i in range(1,1000):
                    idx='A'+str(i)
                    if ws_pr[idx].value==None:
                        ws_pr[idx].value=text[0]
                        break
            else:
                for i in range(1,1000):
                    idx='E'+str(i)
                    if ws_pr[idx].value==None:
                        ws_pr[idx].value=text[0]
                        break
            sms_count+=1
            sms_idx='A'+str(sms_count)
            sms=ws[sms_idx].value
            wb_pr.save('messages.xlsx')
    wb_pr.close()
    wb.close()
mb=openpyxl.load_workbook('messages.xlsx')
ms=mb.active
with st.sidebar:
    selected = option_menu("Select Category", ["HAM", 'Fund Transfer','Login OTPs','Password Reset OTPs','SPAM'],default_index=1)
    selected    
if selected=="HAM":
    for i in range(2,1000):
        idx='A'+str(i)
        if ms[idx].value:
            cell=ms[idx].value
            st.text(cell)
        else:
            break
if selected=="Fund Transfer":
    for i in range(2,1000):
        idx='B'+str(i)
        if ms[idx].value:
            cell=ms[idx].value
            st.text(cell)
        else:
            break
if selected=="Login OTPs":
    for i in range(2,1000):
        idx='C'+str(i)
        if ms[idx].value:
            cell=ms[idx].value
            st.text(cell)
        else:
            break
if selected=="Password Reset OTPs":
    for i in range(2,1000):
        idx='D'+str(i)
        if ms[idx].value:
            cell=ms[idx].value
            st.text(cell)
        else:
            break
if selected=='SPAM':
    for i in range(2,1000):
        idx='E'+str(i)
        if ms[idx].value:
            cell=ms[idx].value
            st.text(cell)
        else:
            break
mb.close()    


                
            
            
                        
                    
                    
                
    
    
    


